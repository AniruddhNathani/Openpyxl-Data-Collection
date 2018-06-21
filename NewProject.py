# -*- coding: utf-8 -*-
"""
Created on Wed Jun 20 14:08:44 2018

@author: nathani_n
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

from openpyxl.utils import coordinate_from_string, column_index_from_string


#NOTE: PROCESS TYPE TESTED CHANGED to PROCESS TYPE TESTED A, PROCESS TYPE TESTED B, PROCESS TYPE TESTED C 
#.......in the original file so as to distinguish the cells A20, A25 and A34

list_search_items=['ModuleId','Change Number','Process Type Tested', 'Project/Problem Details', 'Impacted Application Instance' ,
                   'Impacted Process','Impacted Region']

num=len(list_search_items)
flag=0
for item in list_search_items:
    if item == 'Process Type Tested':
        flag=1
        list_search_items.remove('Process Type Tested')
        list_search_items.append('Process Type Tested')
        break

wb1=Workbook()
dest_filename=r'C:\Users\nathani_n\Desktop\ExcelData\FinalDisplay.xlsx'
ws1 = wb1.active
ws1.title = "Final Data"
if(flag==1):
    list_search_items.extend(['Process Type Tested_RTI','Process Type Tested_NFTI'])
    ws1.append(list_search_items)
else:
    ws1.append(list_search_items)

wb2= Workbook()                
wb2= load_workbook(r'C:\Users\nathani_n\Desktop\ExcelData\Predictive_Analytics_Consolidated_MIM_Information_v1.xlsx')
ws2=wb2[wb2.sheetnames[0]]
xy=[]
row_main=[]
for i in range(1,num):
    for row_i in range(1, ws2.max_row + 1):
        for column_i in range(1,3):
            if (str(ws2.cell(row=row_i, column=column_i).value)).strip() == str(list_search_items[i]).strip():
                row_main.append(row_i)

print(row_main) 
list_new=[]
cell_values_sheet=[]
for sheets in wb2:
    cell_values=[]
    cell_values.append(str(sheets))
    for i in range(len(row_main)):
        for row in sheets.iter_rows('B{}'.format(row_main[i])):
                for col in sheets.iter_cols(min_col=2,max_col=2):
                    for cell in row:
                        cell_values.append(cell.value)
    ws1.append(cell_values)      
    wb1.save(filename = dest_filename)


data_xls = pd.read_excel(r'C:\Users\nathani_n\Desktop\ExcelData\FinalDisplay.xlsx', index_col=None)
data_xls.to_csv(r'C:\Users\nathani_n\Desktop\ExcelData\FINALCSV.csv', index=False, encoding='utf-8')



