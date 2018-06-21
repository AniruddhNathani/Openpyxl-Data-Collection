# -*- coding: utf-8 -*-
"""
Created on Wed Jun 20 14:08:44 2018

@author: nathani_n
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import Counter

from openpyxl.utils import coordinate_from_string, column_index_from_string


list_search_items=['ModuleId','Change Number','Process Type Tested', 'Project/Problem Details', 'Impacted Application Instance' ,
                   'Impacted Process','Impacted Region']

wb2= Workbook()                
wb2= load_workbook(r'C:\Users\nathani_n\Desktop\ExcelData\Predictive_Analytics_Consolidated_MIM_Information_v1.xlsx')
ws2=wb2[wb2.sheetnames[0]]



column1_data=[]
for row in ws2.iter_rows('A{}:A{}'.format(ws2.min_row,ws2.max_row)):
    for col in ws2.iter_cols(min_col=1,max_col=1):
        for cell in row:
            column1_data.append(cell.value)

dictionary_column1_data=dict(Counter(column1_data))
multiple_occurence_data_list=[]
for key,value in dictionary_column1_data.items():
    if (value>1):
        multiple_occurence_data_list.append(key)

wb1=Workbook()
dest_filename=r'C:\Users\nathani_n\Desktop\ExcelData\OutageDf.xlsx'
ws1 = wb1.active
ws1.title = "Outage Data File"

num=len(list_search_items)
k=0
for item in list_search_items:
    if item in multiple_occurence_data_list :
        list_search_items.remove('{}'.format(item))
        list_search_items.append('{}'.format(item))
        list_search_items.extend(['{}{}'.format(item,'_RTI'),'{}{}'.format(item,'_NFTI')])
        k+=1
        print(k)

        if k==len(multiple_occurence_data_list):
            ws1.append(list_search_items)
            break
    
xy=[]
row_main=[]
for i in range(1,num):
    for row_i in range(1, ws2.max_row + 1):
        for column_i in range(1,2):
            if (str(ws2.cell(row=row_i, column=column_i).value)).strip() == str(list_search_items[i]).strip():
                row_main.append(row_i)
                
list_new=[]
cell_values_sheet=[]
for sheets in wb2:
    cell_values=[]
    cell_values.append(str(sheets))
    for i in range(len(row_main)):
        for row in sheets.iter_rows('B{}'.format(row_main[i])):
                for col in sheets.iter_cols(min_col=2,max_col=2):
                    for cell in row:
                        cell_values.append(cell.value)
    ws1.append(cell_values)      
    wb1.save(filename = dest_filename)


data_xls = pd.read_excel(r'C:\Users\nathani_n\Desktop\ExcelData\OutageDf.xlsx', index_col=None)
data_xls.to_csv(r'C:\Users\nathani_n\Desktop\ExcelData\OutageDfCSV.csv', index=False, encoding='utf-8')




